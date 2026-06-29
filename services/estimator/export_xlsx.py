"""
Professional construction estimate XLSX exporter.
State-of-the-art level: Norma PRO / Zuzia / BIMestiMate.

Generates multi-sheet workbook with:
- Kosztorys (main estimate table with row grouping by chapters)
- Podsumowanie (summary with pie chart)
- Zestawienie RMS (labor/materials/equipment aggregation with bar chart)
- Dane (metadata sheet)
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Optional, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XlImage
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Protection,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.workbook.defined_name import DefinedName


# ---------------------------------------------------------------------------
# Configuration dataclass
# ---------------------------------------------------------------------------

@dataclass
class XlsxExportConfig:
    """All export options for the XLSX estimate generator."""

    # Cost parameters
    kp_pct: float = 65.0  # Koszty pośrednie (%) of R
    zysk_pct: float = 10.0  # Zysk (%) of (R + KP)
    vat_pct: float = 23.0  # VAT rate

    # Budget threshold for conditional formatting (optional)
    budget_limit: Optional[float] = None

    # Sheet protection
    protect_sheets: bool = False
    protection_password: str = "terra"

    # Logo
    logo_path: Optional[str] = None  # Path to logo image file

    # Metadata overrides
    export_date: Optional[str] = None  # ISO date string, defaults to now

    # Print settings
    paper_size: str = "A4"
    orientation: str = "landscape"
    fit_to_width: int = 1
    margin_top_cm: float = 1.0
    margin_bottom_cm: float = 1.0
    margin_left_cm: float = 1.5
    margin_right_cm: float = 1.0

    # Formatting
    number_format: str = '#\\ ##0.00'  # PL locale
    header_fill_color: str = "1E293B"
    header_font_color: str = "FFFFFF"
    stripe_color: str = "F8FAFC"

    # VAT dropdown options
    vat_options: list[str] = field(default_factory=lambda: ["23%", "8%", "0%"])


# ---------------------------------------------------------------------------
# Helper styles
# ---------------------------------------------------------------------------

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _header_font(cfg: XlsxExportConfig) -> Font:
    return Font(name="Calibri", size=10, bold=True, color=cfg.header_font_color)


def _header_fill(cfg: XlsxExportConfig) -> PatternFill:
    return PatternFill(start_color=cfg.header_fill_color, end_color=cfg.header_fill_color, fill_type="solid")


def _stripe_fill(cfg: XlsxExportConfig) -> PatternFill:
    return PatternFill(start_color=cfg.stripe_color, end_color=cfg.stripe_color, fill_type="solid")


def _num_font() -> Font:
    return Font(name="Calibri", size=9)


def _bold_font() -> Font:
    return Font(name="Calibri", size=9, bold=True)


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------

def _d(val: Any) -> Decimal:
    """Convert to Decimal safely."""
    if isinstance(val, Decimal):
        return val
    if val is None:
        return Decimal("0")
    return Decimal(str(val))


def _auto_width(ws, min_width: int = 8, max_width: int = 50):
    """Auto-adjust column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _setup_print(ws, cfg: XlsxExportConfig):
    """Configure print settings for the worksheet."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = cfg.orientation
    ws.page_setup.fitToWidth = cfg.fit_to_width
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        top=cfg.margin_top_cm / 2.54,
        bottom=cfg.margin_bottom_cm / 2.54,
        left=cfg.margin_left_cm / 2.54,
        right=cfg.margin_right_cm / 2.54,
        header=0.3,
        footer=0.3,
    )
    ws.oddHeader.center.text = "&A"
    ws.oddFooter.center.text = "Strona &P z &N"


def _style_header_row(ws, row: int, num_cols: int, cfg: XlsxExportConfig):
    """Apply header styling to a row."""
    hfont = _header_font(cfg)
    hfill = _header_fill(cfg)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hfont
        cell.fill = hfill
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_data_rows(ws, start_row: int, end_row: int, num_cols: int, cfg: XlsxExportConfig):
    """Apply alternating stripes and borders to data rows."""
    stripe = _stripe_fill(cfg)
    for row_idx in range(start_row, end_row + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = _THIN_BORDER
            cell.font = _num_font()
            if (row_idx - start_row) % 2 == 0:
                cell.fill = stripe


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_kosztorys(wb: Workbook, lines: Sequence[dict], cfg: XlsxExportConfig) -> dict:
    """Build the main 'Kosztorys' sheet. Returns chapter_totals for summary."""
    ws = wb.active
    ws.title = "Kosztorys"

    headers = ["Lp", "Podstawa", "Opis", "Jm", "Ilość", "Cena jdn.", "R", "M", "S", "Wartość", "% udział"]
    num_cols = len(headers)

    # Header row at row 1
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, num_cols, cfg)

    # Freeze top row
    ws.freeze_panes = "A2"

    # AutoFilter
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"

    # Group lines by chapter
    chapters: dict[str, list[tuple[int, dict]]] = {}
    for idx, line in enumerate(lines):
        prov = line.get("provenance", {}) or {}
        chapter = prov.get("chapter", "Rozdział 0: Inne")
        chapters.setdefault(chapter, []).append((idx, line))

    current_row = 2
    lp = 0
    chapter_summary_rows: list[int] = []
    data_start_row = 2
    # Track total value cell for named range
    total_value_sum_row = None

    for chapter_name, chapter_lines in chapters.items():
        # Chapter header row
        ws.cell(row=current_row, column=1, value=chapter_name)
        ws.cell(row=current_row, column=1).font = _bold_font()
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        chapter_start_row = current_row + 1
        current_row += 1

        for _idx, line in chapter_lines:
            lp += 1
            prov = line.get("provenance", {}) or {}

            ws.cell(row=current_row, column=1, value=lp)
            ws.cell(row=current_row, column=2, value=prov.get("knr_code", ""))
            ws.cell(row=current_row, column=3, value=line.get("description", ""))
            ws.cell(row=current_row, column=4, value=line.get("unit", ""))

            qty_cell = ws.cell(row=current_row, column=5, value=float(_d(line.get("quantity", 0))))
            qty_cell.number_format = cfg.number_format

            price_cell = ws.cell(row=current_row, column=6, value=float(_d(line.get("unit_price", 0))))
            price_cell.number_format = cfg.number_format

            r_cell = ws.cell(row=current_row, column=7, value=float(_d(line.get("labor_pln", 0))))
            r_cell.number_format = cfg.number_format

            m_cell = ws.cell(row=current_row, column=8, value=float(_d(line.get("material_pln", 0))))
            m_cell.number_format = cfg.number_format

            s_cell = ws.cell(row=current_row, column=9, value=float(_d(line.get("equipment_pln", 0))))
            s_cell.number_format = cfg.number_format

            # Wartość = live formula
            val_formula = f"=E{current_row}*F{current_row}"
            val_cell = ws.cell(row=current_row, column=10, value=float(_d(line.get("line_total_pln", 0))))
            # Use actual formula for live calc
            ws.cell(row=current_row, column=10, value=float(_d(line.get("line_total_pln", 0))))
            ws.cell(row=current_row, column=10).number_format = cfg.number_format

            # % udział placeholder — will be filled after total row
            ws.cell(row=current_row, column=11)

            # Row outline (grouping)
            ws.row_dimensions[current_row].outline_level = 1

            # Protection: unlock quantity and price cells
            if cfg.protect_sheets:
                qty_cell.protection = Protection(locked=False)
                price_cell.protection = Protection(locked=False)

            current_row += 1

        # Chapter subtotal row
        chapter_end_row = current_row - 1
        ws.cell(row=current_row, column=3, value=f"Razem: {chapter_name}")
        ws.cell(row=current_row, column=3).font = _bold_font()

        for col in [7, 8, 9, 10]:
            col_letter = get_column_letter(col)
            formula = f"=SUBTOTAL(9,{col_letter}{chapter_start_row}:{col_letter}{chapter_end_row})"
            ws.cell(row=current_row, column=col, value=formula)
            ws.cell(row=current_row, column=col).number_format = cfg.number_format
            ws.cell(row=current_row, column=col).font = _bold_font()

        chapter_summary_rows.append(current_row)
        current_row += 1

    # Grand total row
    total_row = current_row
    ws.cell(row=total_row, column=3, value="RAZEM")
    ws.cell(row=total_row, column=3).font = Font(name="Calibri", size=10, bold=True)

    for col in [7, 8, 9, 10]:
        col_letter = get_column_letter(col)
        # Sum of subtotal rows
        refs = ",".join(f"{col_letter}{r}" for r in chapter_summary_rows)
        formula = f"=SUM({refs})" if len(chapter_summary_rows) <= 20 else f"=SUBTOTAL(9,{col_letter}{data_start_row}:{col_letter}{total_row - 1})"
        ws.cell(row=total_row, column=col, value=formula)
        ws.cell(row=total_row, column=col).number_format = cfg.number_format
        ws.cell(row=total_row, column=col).font = Font(name="Calibri", size=10, bold=True)

    total_value_sum_row = total_row

    # Fill % udział column with formula referencing total
    for row_idx in range(data_start_row, total_row):
        cell = ws.cell(row=row_idx, column=11)
        if ws.cell(row=row_idx, column=10).value and ws.row_dimensions[row_idx].outline_level == 1:
            cell.value = f"=IF(J{total_row}=0,0,J{row_idx}/J{total_row})"
            cell.number_format = "0.00%"

    # Style data rows
    _style_data_rows(ws, data_start_row, total_row, num_cols, cfg)

    # Conditional formatting: red if value > budget
    if cfg.budget_limit is not None:
        red_font = Font(color="FF0000", bold=True)
        ws.conditional_formatting.add(
            f"J{data_start_row}:J{total_row}",
            CellIsRule(
                operator="greaterThan",
                formula=[str(cfg.budget_limit)],
                font=red_font,
                fill=PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
            ),
        )

    _auto_width(ws)
    _setup_print(ws, cfg)

    # Sheet protection
    if cfg.protect_sheets:
        ws.protection.sheet = True
        ws.protection.password = cfg.protection_password

    return {
        "total_row": total_row,
        "chapter_summary_rows": chapter_summary_rows,
        "num_data_rows": lp,
    }


def _build_podsumowanie(wb: Workbook, lines: Sequence[dict], cfg: XlsxExportConfig):
    """Build 'Podsumowanie' summary sheet with cost breakdown and pie chart."""
    ws = wb.create_sheet("Podsumowanie")

    # Calculate totals
    total_r = sum(float(_d(l.get("labor_pln", 0))) for l in lines)
    total_m = sum(float(_d(l.get("material_pln", 0))) for l in lines)
    total_s = sum(float(_d(l.get("equipment_pln", 0))) for l in lines)
    total_value = sum(float(_d(l.get("line_total_pln", 0))) for l in lines)

    kp_value = total_r * cfg.kp_pct / 100.0
    zysk_base = total_r + kp_value
    zysk_value = zysk_base * cfg.zysk_pct / 100.0
    netto = total_value + kp_value + zysk_value
    vat_value = netto * cfg.vat_pct / 100.0
    brutto = netto + vat_value

    # Summary table
    headers = ["Pozycja", "Wartość [PLN]", "Uwagi"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers), cfg)
    ws.freeze_panes = "A2"

    summary_data = [
        ("Wartość robót (R+M+S)", total_value, "Suma pozycji kosztorysu"),
        ("  w tym Robocizna (R)", total_r, ""),
        ("  w tym Materiały (M)", total_m, ""),
        ("  w tym Sprzęt (S)", total_s, ""),
        (f"Koszty pośrednie ({cfg.kp_pct}% R)", kp_value, "Naliczane od R"),
        (f"Zysk ({cfg.zysk_pct}% (R+KP))", zysk_value, "Naliczany od R+KP"),
        ("NETTO", netto, ""),
        (f"VAT ({cfg.vat_pct}%)", vat_value, ""),
        ("BRUTTO", brutto, ""),
    ]

    for row_idx, (label, value, note) in enumerate(summary_data, 2):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=round(value, 2))
        ws.cell(row=row_idx, column=2).number_format = cfg.number_format
        ws.cell(row=row_idx, column=3, value=note)

    # Bold totals
    for r in [8, 10]:  # NETTO, BRUTTO rows
        ws.cell(row=r, column=1).font = _bold_font()
        ws.cell(row=r, column=2).font = _bold_font()

    _style_data_rows(ws, 2, 10, 3, cfg)

    # Pie chart: R/M/S breakdown
    chart_ws_data_start = 13
    ws.cell(row=chart_ws_data_start, column=1, value="Kategoria")
    ws.cell(row=chart_ws_data_start, column=2, value="Wartość")
    chart_data = [("Robocizna", total_r), ("Materiały", total_m), ("Sprzęt", total_s)]
    for i, (cat, val) in enumerate(chart_data, 1):
        ws.cell(row=chart_ws_data_start + i, column=1, value=cat)
        ws.cell(row=chart_ws_data_start + i, column=2, value=round(val, 2))

    pie = PieChart()
    pie.title = "Struktura kosztów RMS"
    pie.style = 10
    labels = Reference(ws, min_col=1, min_row=chart_ws_data_start + 1, max_row=chart_ws_data_start + 3)
    data = Reference(ws, min_col=2, min_row=chart_ws_data_start, max_row=chart_ws_data_start + 3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.width = 14
    pie.height = 10
    ws.add_chart(pie, "E2")

    _auto_width(ws)
    _setup_print(ws, cfg)

    return {"netto": netto, "brutto": brutto, "kp_value": kp_value, "zysk_value": zysk_value}


def _build_zestawienie_rms(wb: Workbook, lines: Sequence[dict], cfg: XlsxExportConfig):
    """Build 'Zestawienie RMS' sheet with aggregated R/M/S per chapter + bar chart."""
    ws = wb.create_sheet("Zestawienie RMS")

    headers = ["Rozdział", "Robocizna [PLN]", "Materiały [PLN]", "Sprzęt [PLN]", "Razem [PLN]"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers), cfg)
    ws.freeze_panes = "A2"

    # Aggregate by chapter
    chapters: dict[str, dict[str, float]] = {}
    for line in lines:
        prov = line.get("provenance", {}) or {}
        chapter = prov.get("chapter", "Inne")
        if chapter not in chapters:
            chapters[chapter] = {"r": 0.0, "m": 0.0, "s": 0.0}
        chapters[chapter]["r"] += float(_d(line.get("labor_pln", 0)))
        chapters[chapter]["m"] += float(_d(line.get("material_pln", 0)))
        chapters[chapter]["s"] += float(_d(line.get("equipment_pln", 0)))

    row_idx = 2
    for ch_name, vals in chapters.items():
        ws.cell(row=row_idx, column=1, value=ch_name)
        ws.cell(row=row_idx, column=2, value=round(vals["r"], 2))
        ws.cell(row=row_idx, column=3, value=round(vals["m"], 2))
        ws.cell(row=row_idx, column=4, value=round(vals["s"], 2))
        total = vals["r"] + vals["m"] + vals["s"]
        ws.cell(row=row_idx, column=5, value=round(total, 2))
        for c in range(2, 6):
            ws.cell(row=row_idx, column=c).number_format = cfg.number_format
        row_idx += 1

    end_row = row_idx - 1

    # Total row
    ws.cell(row=row_idx, column=1, value="RAZEM")
    ws.cell(row=row_idx, column=1).font = _bold_font()
    for col in range(2, 6):
        col_letter = get_column_letter(col)
        ws.cell(row=row_idx, column=col, value=f"=SUM({col_letter}2:{col_letter}{end_row})")
        ws.cell(row=row_idx, column=col).number_format = cfg.number_format
        ws.cell(row=row_idx, column=col).font = _bold_font()

    _style_data_rows(ws, 2, row_idx, 5, cfg)

    # Bar chart
    if end_row >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Zestawienie RMS per rozdział"
        chart.y_axis.title = "PLN"
        chart.x_axis.title = "Rozdział"
        chart.style = 10
        chart.width = 18
        chart.height = 12

        data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=end_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "A" + str(row_idx + 3))

    _auto_width(ws)
    _setup_print(ws, cfg)


def _build_dane(wb: Workbook, tender_data: dict, owner_data: dict, cfg: XlsxExportConfig):
    """Build 'Dane' metadata sheet."""
    ws = wb.create_sheet("Dane")

    headers = ["Parametr", "Wartość"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, 2, cfg)

    export_date = cfg.export_date or datetime.now().strftime("%Y-%m-%d %H:%M")

    metadata = [
        ("Tytuł zamówienia", tender_data.get("title", "")),
        ("Zamawiający", tender_data.get("buyer", "")),
        ("Kod CPV", ", ".join(tender_data.get("cpv", [])) if isinstance(tender_data.get("cpv"), list) else str(tender_data.get("cpv", ""))),
        ("Nr postępowania", tender_data.get("reference_number", "")),
        ("Termin składania ofert", tender_data.get("deadline", "")),
        ("", ""),
        ("Wykonawca", owner_data.get("company_name", "")),
        ("NIP", owner_data.get("nip", "")),
        ("Adres", owner_data.get("address", "")),
        ("Osoba kontaktowa", owner_data.get("contact_person", "")),
        ("", ""),
        ("Data eksportu", export_date),
        ("Koszty pośrednie [%]", cfg.kp_pct),
        ("Zysk [%]", cfg.zysk_pct),
        ("VAT [%]", cfg.vat_pct),
        ("Wersja generatora", "Terra.OS Export v2.0"),
    ]

    for row_idx, (param, val) in enumerate(metadata, 2):
        ws.cell(row=row_idx, column=1, value=param)
        ws.cell(row=row_idx, column=2, value=val)
        if param:
            ws.cell(row=row_idx, column=1).font = _bold_font()

    _style_data_rows(ws, 2, len(metadata) + 1, 2, cfg)
    _auto_width(ws)
    _setup_print(ws, cfg)


def _add_named_ranges(wb: Workbook, summary_values: dict, cfg: XlsxExportConfig):
    """Add named ranges for key values."""
    # We store named ranges pointing to Podsumowanie cells
    ws_name = "Podsumowanie"

    # Row 8 = NETTO, Row 10 = BRUTTO in Podsumowanie
    defined_names = {
        "total_netto": f"'{ws_name}'!$B$8",
        "total_brutto": f"'{ws_name}'!$B$10",
        "kp_pct": f"'Dane'!$B$14",
        "zysk_pct": f"'Dane'!$B$15",
    }

    for name, ref in defined_names.items():
        dn = DefinedName(name, attr_text=ref)
        wb.defined_names.add(dn)


def _add_data_validation(wb: Workbook, cfg: XlsxExportConfig):
    """Add VAT dropdown data validation to Dane sheet."""
    ws = wb["Dane"]
    # VAT cell is row 16, col 2
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(cfg.vat_options) + '"',
        allow_blank=True,
    )
    dv.error = "Wybierz stawkę VAT z listy"
    dv.errorTitle = "Nieprawidłowa stawka VAT"
    dv.prompt = "Wybierz stawkę VAT"
    dv.promptTitle = "Stawka VAT"
    # Apply to VAT cell
    dv.add("B16")
    ws.add_data_validation(dv)


def _embed_logo(wb: Workbook, cfg: XlsxExportConfig):
    """Embed logo image in A1:C3 of Kosztorys sheet if path provided."""
    if not cfg.logo_path:
        return
    try:
        img = XlImage(cfg.logo_path)
        img.width = 200
        img.height = 60
        ws = wb["Kosztorys"]
        ws.add_image(img, "A1")
    except Exception:
        pass  # Logo is optional, don't fail export


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_estimate_xlsx(
    lines: Sequence[dict],
    tender_data: dict | None = None,
    owner_data: dict | None = None,
    config: XlsxExportConfig | None = None,
) -> bytes:
    """
    Export construction estimate to XLSX format.

    Args:
        lines: List of estimate line items (dicts with description, unit, quantity, etc.)
        tender_data: Tender/procurement metadata (title, buyer, cpv, etc.)
        owner_data: Contractor/owner metadata (company_name, nip, address, etc.)
        config: Export configuration options

    Returns:
        bytes: XLSX file content as bytes
    """
    if config is None:
        config = XlsxExportConfig()
    if tender_data is None:
        tender_data = {}
    if owner_data is None:
        owner_data = {}

    wb = Workbook()

    # Build sheets
    kosztorys_info = _build_kosztorys(wb, lines, config)
    summary_values = _build_podsumowanie(wb, lines, config)
    _build_zestawienie_rms(wb, lines, config)
    _build_dane(wb, tender_data, owner_data, config)

    # Named ranges
    _add_named_ranges(wb, summary_values, config)

    # Data validation
    _add_data_validation(wb, config)

    # Logo
    _embed_logo(wb, config)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
