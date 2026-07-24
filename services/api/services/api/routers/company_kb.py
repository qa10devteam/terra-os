"""Baza Wiedzy Firmy — company knowledge base.

Moduł pozwala firmie budowlanej gromadzić:
  - profil firmy (NIP, certyfikaty, personel, kontekst AI)
  - własne stawki R/M/S (import Excel lub ręcznie)
  - referencje projektów (AI generuje streszczenia do ofert)
  - historia przetargów (nasze ceny vs wynik vs rynek)

Endpoints:
  GET/PUT  /api/v2/company/profile           — profil + kontekst AI
  GET      /api/v2/company/profile/ai-context — kontekst do użycia przez silnik ofert
  GET/POST /api/v2/company/rates              — własne stawki
  POST     /api/v2/company/rates/import-excel — import z Excela (xlsx)
  GET/POST /api/v2/company/references         — referencje projektów
  PUT      /api/v2/company/references/{id}    — aktualizuj referencję
  DELETE   /api/v2/company/references/{id}    — usuń referencję
  GET      /api/v2/company/bids               — historia przetargów (historical_bids)
  POST     /api/v2/company/bids               — dodaj wpis historyczny
  GET      /api/v2/company/bids/stats         — statystyki win-rate, avg margin
"""
from __future__ import annotations

import json
import logging
import uuid
from datetime import datetime, date
from typing import Any, Optional

import sqlalchemy as sa
from fastapi import APIRouter, File, HTTPException, UploadFile
from pydantic import BaseModel

from terra_db.session import get_engine
from ..auth.deps import AuthUser

router = APIRouter(prefix="/api/v2/company", tags=["company-kb"])
logger = logging.getLogger(__name__)


# ─── Schematy wejściowe ────────────────────────────────────────────────────────

class ProfileUpdate(BaseModel):
    company_name: Optional[str] = None
    nip: Optional[str] = None
    regon: Optional[str] = None
    krs: Optional[str] = None
    adres: Optional[str] = None
    uprawnienia: Optional[list[str]] = None
    personel_kluczowy: Optional[list[dict]] = None
    certyfikaty: Optional[list[str]] = None
    cpv_preferred: Optional[list[str]] = None
    voivodeships: Optional[list[str]] = None
    scope_notes: Optional[str] = None
    ai_context_md: Optional[str] = None  # ręczny kontekst AI
    rate_card: Optional[dict] = None     # KP%, KZ%, zysk%, robocizna_zl_rg


class RateIn(BaseModel):
    symbol: str
    nazwa: str
    jednostka: Optional[str] = "rg"
    typ_rms: str  # R / M / S
    cena_netto: float
    katalog: Optional[str] = None
    uwagi: Optional[str] = None


class ReferenceIn(BaseModel):
    nazwa: str
    inwestor: Optional[str] = None
    lokalizacja: Optional[str] = None
    rok_realizacji: Optional[int] = None
    wartosc_pln: Optional[float] = None
    cpv_codes: Optional[list[str]] = []
    zakres_md: Optional[str] = None
    certyfikaty: Optional[list[str]] = []
    source_doc_id: Optional[str] = None


class BidIn(BaseModel):
    tender_id: Optional[str] = None
    cpv: Optional[str] = None
    region: Optional[str] = None
    our_price: float
    winning_price: Optional[float] = None
    n_competitors: Optional[int] = None
    won: bool = False
    markup_pct: Optional[float] = None
    actual_cost: Optional[float] = None
    margin_pct: Optional[float] = None
    bid_date: Optional[str] = None  # ISO date string


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _tenant(user: AuthUser) -> str:
    return str(user.org_id or user.user_id)


def _get_profile(conn, tenant_id: str) -> dict | None:
    row = conn.execute(
        sa.text("SELECT * FROM owner_profile WHERE tenant_id = :tid LIMIT 1"),
        {"tid": tenant_id},
    ).mappings().fetchone()
    return dict(row) if row else None


def _build_ai_context(profile: dict, rates: list[dict], refs: list[dict], bids_stats: dict) -> str:
    """Buduje blok kontekstu dla silnika AI ofert."""
    lines = ["# Baza Wiedzy Firmy — kontekst dla AI\n"]

    # Dane firmy
    lines.append(f"## Firma\n- Nazwa: {profile.get('company_name', '—')}")
    if profile.get("nip"):     lines.append(f"- NIP: {profile['nip']}")
    if profile.get("adres"):   lines.append(f"- Adres: {profile['adres']}")
    if profile.get("uprawnienia"):
        lines.append(f"- Uprawnienia: {', '.join(profile['uprawnienia'])}")
    if profile.get("certyfikaty"):
        lines.append(f"- Certyfikaty: {', '.join(profile['certyfikaty'])}")
    if profile.get("scope_notes"):
        lines.append(f"- Specjalizacja: {profile['scope_notes']}")

    # Personel kluczowy
    personel = profile.get("personel_kluczowy") or []
    if personel:
        lines.append("\n## Personel kluczowy")
        for p in personel[:10]:
            lines.append(f"- {p.get('imie_nazwisko','?')} — {p.get('rola','?')} {p.get('uprawnienia','')}")

    # Własne stawki R (robocizna)
    if rates:
        r_rates = [r for r in rates if r.get("typ_rms") == "R"]
        if r_rates:
            avg_r = sum(r["cena_netto"] for r in r_rates) / len(r_rates)
            lines.append(f"\n## Stawki własne\n- Robocizna (śr.): {avg_r:.2f} PLN/rg")
        m_rates = [r for r in rates if r.get("typ_rms") == "M"]
        if m_rates:
            lines.append(f"- Materiały (pozycji): {len(m_rates)}")

    # Rate card (narzuty)
    rc = profile.get("rate_card") or {}
    if rc:
        lines.append(f"- KP%: {rc.get('kp_pct','?')}  KZ%: {rc.get('kz_pct','?')}  Zysk%: {rc.get('zysk_pct','?')}")

    # Referencje
    if refs:
        lines.append(f"\n## Referencje ({len(refs)} projektów)")
        for ref in refs[:8]:
            val = f"{ref['wartosc_pln']:,.0f} PLN" if ref.get("wartosc_pln") else "—"
            lines.append(f"- **{ref['nazwa']}** ({ref.get('rok_realizacji','?')}) — {ref.get('inwestor','—')} — {val}")
            if ref.get("zakres_md"):
                lines.append(f"  > {ref['zakres_md'][:150]}")

    # Historia ofert
    if bids_stats.get("total"):
        lines.append(f"\n## Historia przetargów")
        lines.append(f"- Złożone oferty: {bids_stats['total']}")
        lines.append(f"- Win-rate: {bids_stats.get('win_rate_pct', 0):.1f}%")
        lines.append(f"- Średni narzut: {bids_stats.get('avg_markup_pct', 0):.1f}%")
        if bids_stats.get("avg_margin_pct"):
            lines.append(f"- Śr. marża (faktyczna): {bids_stats['avg_margin_pct']:.1f}%")

    # Ręczny kontekst AI
    if profile.get("ai_context_md"):
        lines.append(f"\n## Dodatkowe informacje\n{profile['ai_context_md']}")

    return "\n".join(lines)


# ─── PROFILE ──────────────────────────────────────────────────────────────────

@router.get("/profile")
def get_profile(user: AuthUser) -> dict[str, Any]:
    """Profil firmy z pełnym kontekstem."""
    engine = get_engine()
    tid = _tenant(user)
    with engine.connect() as conn:
        profile = _get_profile(conn, tid)
    if not profile:
        return {"company_name": "", "nip": "", "adres": "", "rate_card": {},
                "uprawnienia": [], "certyfikaty": [], "cpv_preferred": [],
                "voivodeships": [], "scope_notes": "", "ai_context_md": "",
                "personel_kluczowy": []}
    # Konwersja typów JSONB/ARRAY
    profile["rate_card"]       = profile.get("rate_card") or {}
    profile["personel_kluczowy"] = profile.get("personel_kluczowy") or []
    return dict(profile)


@router.put("/profile")
def update_profile(data: ProfileUpdate, user: AuthUser) -> dict[str, Any]:
    """Zaktualizuj profil firmy."""
    engine = get_engine()
    tid = _tenant(user)
    updates: dict[str, Any] = {k: v for k, v in data.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.utcnow()

    set_clause = ", ".join(f"{k} = :{k}" for k in updates)
    updates["tid"] = tid

    with engine.begin() as conn:
        existing = _get_profile(conn, tid)
        if existing:
            conn.execute(
                sa.text(f"UPDATE owner_profile SET {set_clause} WHERE tenant_id = :tid"),
                updates,
            )
        else:
            updates["id"] = str(uuid.uuid4())
            cols = ", ".join(updates.keys())
            vals = ", ".join(f":{k}" for k in updates)
            conn.execute(sa.text(f"INSERT INTO owner_profile ({cols}) VALUES ({vals})"), updates)

    return {"ok": True}


@router.get("/profile/ai-context")
def get_ai_context(user: AuthUser) -> dict[str, str]:
    """Zwraca gotowy blok kontekstu AI do użycia przez silnik ofert."""
    engine = get_engine()
    tid = _tenant(user)
    with engine.connect() as conn:
        profile = _get_profile(conn, tid) or {}
        rates = conn.execute(
            sa.text("SELECT typ_rms, cena_netto, nazwa FROM company_rates_import WHERE tenant_id=:tid AND aktywna=TRUE LIMIT 100"),
            {"tid": tid},
        ).mappings().fetchall()
        refs = conn.execute(
            sa.text("SELECT nazwa, inwestor, rok_realizacji, wartosc_pln, zakres_md FROM company_references WHERE tenant_id=:tid ORDER BY rok_realizacji DESC NULLS LAST LIMIT 20"),
            {"tid": tid},
        ).mappings().fetchall()
        stats_row = conn.execute(
            sa.text("""SELECT COUNT(*) total, SUM(CASE WHEN won THEN 1 ELSE 0 END) wins,
                              AVG(markup_pct) avg_markup, AVG(margin_pct) avg_margin
                       FROM historical_bids WHERE org_id=:tid"""),
            {"tid": tid},
        ).fetchone()

    bids_stats = {}
    if stats_row and stats_row[0]:
        bids_stats = {
            "total": stats_row[0],
            "win_rate_pct": (stats_row[1] or 0) / stats_row[0] * 100,
            "avg_markup_pct": float(stats_row[2] or 0),
            "avg_margin_pct": float(stats_row[3] or 0) if stats_row[3] else None,
        }

    context = _build_ai_context(
        dict(profile),
        [dict(r) for r in rates],
        [dict(r) for r in refs],
        bids_stats,
    )
    return {"context": context, "chars": str(len(context))}


# ─── RATES ────────────────────────────────────────────────────────────────────

@router.get("/rates")
def list_rates(user: AuthUser) -> dict[str, Any]:
    engine = get_engine()
    tid = _tenant(user)
    with engine.connect() as conn:
        rows = conn.execute(
            sa.text("""SELECT id, symbol, nazwa, jednostka, typ_rms, cena_netto,
                              katalog, uwagi, source, created_at
                       FROM company_rates_import
                       WHERE tenant_id=:tid AND aktywna=TRUE
                       ORDER BY typ_rms, nazwa"""),
            {"tid": tid},
        ).mappings().fetchall()
    return {"items": [dict(r) for r in rows], "total": len(rows)}


@router.post("/rates")
def add_rate(data: RateIn, user: AuthUser) -> dict[str, Any]:
    engine = get_engine()
    tid = _tenant(user)
    rid = str(uuid.uuid4())
    with engine.begin() as conn:
        conn.execute(sa.text("""
            INSERT INTO company_rates_import
                (id, tenant_id, symbol, nazwa, jednostka, typ_rms, cena_netto, katalog, uwagi, source)
            VALUES (:id, :tid, :symbol, :nazwa, :jednostka, :typ_rms, :cena_netto, :katalog, :uwagi, 'manual')
        """), {"id": rid, "tid": tid, **data.model_dump()})
    return {"id": rid, "ok": True}


@router.post("/rates/import-excel")
async def import_rates_excel(user: AuthUser, file: UploadFile = File(...)) -> dict[str, Any]:
    """Import stawek z Excela.

    Oczekiwany format kolumn (dowolna kolejność):
      symbol | nazwa | jednostka | typ_rms | cena_netto | katalog | uwagi
    """
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Wymagany plik Excel (.xlsx lub .xls)")

    try:
        import openpyxl
        content = await file.read()
        import io
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except ImportError:
        raise HTTPException(500, "openpyxl nie zainstalowane — kontakt z adminem")
    except Exception as e:
        raise HTTPException(400, f"Błąd odczytu pliku: {e}")

    # Rozpoznaj nagłówki
    headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {h: i for i, h in enumerate(headers)}

    def _get(row, name: str, default=None):
        idx = col_map.get(name)
        return row[idx].value if idx is not None else default

    engine = get_engine()
    tid = _tenant(user)
    batch_id = str(uuid.uuid4())
    imported = 0
    errors = []

    with engine.begin() as conn:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            try:
                symbol = str(_get(row, "symbol") or "").strip()
                nazwa  = str(_get(row, "nazwa") or "").strip()
                if not symbol or not nazwa:
                    continue
                cena_raw = _get(row, "cena_netto")
                if cena_raw is None:
                    cena_raw = _get(row, "cena")
                try:
                    cena = float(str(cena_raw or "0").replace(",", "."))
                except (ValueError, TypeError):
                    cena = 0.0
                typ = str(_get(row, "typ_rms") or "R").upper()[:1]

                conn.execute(sa.text("""
                    INSERT INTO company_rates_import
                        (id, tenant_id, symbol, nazwa, jednostka, typ_rms, cena_netto,
                         katalog, uwagi, source, import_batch)
                    VALUES (:id, :tid, :symbol, :nazwa, :jednostka, :typ, :cena,
                            :katalog, :uwagi, 'excel_import', :batch)
                    ON CONFLICT DO NOTHING
                """), {
                    "id": str(uuid.uuid4()), "tid": tid,
                    "symbol": symbol, "nazwa": nazwa,
                    "jednostka": str(_get(row, "jednostka") or "rg"),
                    "typ": typ, "cena": cena,
                    "katalog": str(_get(row, "katalog") or ""),
                    "uwagi": str(_get(row, "uwagi") or ""),
                    "batch": batch_id,
                })
                imported += 1
            except Exception as e:
                errors.append(f"Wiersz {row_idx}: {e}")

    return {"imported": imported, "batch_id": batch_id, "errors": errors[:10]}


# ─── REFERENCES ───────────────────────────────────────────────────────────────

@router.get("/references")
def list_references(user: AuthUser) -> dict[str, Any]:
    engine = get_engine()
    tid = _tenant(user)
    with engine.connect() as conn:
        rows = conn.execute(
            sa.text("""SELECT id, nazwa, inwestor, lokalizacja, rok_realizacji,
                              wartosc_pln, cpv_codes, zakres_md, certyfikaty, ai_summary
                       FROM company_references WHERE tenant_id=:tid
                       ORDER BY rok_realizacji DESC NULLS LAST, created_at DESC"""),
            {"tid": tid},
        ).mappings().fetchall()
    return {"items": [dict(r) for r in rows], "total": len(rows)}


@router.post("/references")
def add_reference(data: ReferenceIn, user: AuthUser) -> dict[str, Any]:
    engine = get_engine()
    tid = _tenant(user)
    rid = str(uuid.uuid4())
    with engine.begin() as conn:
        conn.execute(sa.text("""
            INSERT INTO company_references
                (id, tenant_id, nazwa, inwestor, lokalizacja, rok_realizacji,
                 wartosc_pln, cpv_codes, zakres_md, certyfikaty, source_doc_id)
            VALUES (:id, :tid, :nazwa, :inwestor, :lokalizacja, :rok,
                    :wartosc, :cpv, :zakres, :certs, :doc_id)
        """), {
            "id": rid, "tid": tid,
            "nazwa": data.nazwa, "inwestor": data.inwestor,
            "lokalizacja": data.lokalizacja, "rok": data.rok_realizacji,
            "wartosc": data.wartosc_pln,
            "cpv": data.cpv_codes, "zakres": data.zakres_md,
            "certs": data.certyfikaty, "doc_id": data.source_doc_id,
        })
    return {"id": rid, "ok": True}


@router.put("/references/{ref_id}")
def update_reference(ref_id: str, data: ReferenceIn, user: AuthUser) -> dict[str, Any]:
    engine = get_engine()
    tid = _tenant(user)
    with engine.begin() as conn:
        conn.execute(sa.text("""
            UPDATE company_references SET
                nazwa=:nazwa, inwestor=:inwestor, lokalizacja=:lokalizacja,
                rok_realizacji=:rok, wartosc_pln=:wartosc, cpv_codes=:cpv,
                zakres_md=:zakres, certyfikaty=:certs, updated_at=NOW()
            WHERE id=:id AND tenant_id=:tid
        """), {
            "id": ref_id, "tid": tid,
            "nazwa": data.nazwa, "inwestor": data.inwestor,
            "lokalizacja": data.lokalizacja, "rok": data.rok_realizacji,
            "wartosc": data.wartosc_pln,
            "cpv": data.cpv_codes, "zakres": data.zakres_md,
            "certs": data.certyfikaty,
        })
    return {"ok": True}


@router.delete("/references/{ref_id}")
def delete_reference(ref_id: str, user: AuthUser) -> dict[str, Any]:
    engine = get_engine()
    tid = _tenant(user)
    with engine.begin() as conn:
        conn.execute(
            sa.text("DELETE FROM company_references WHERE id=:id AND tenant_id=:tid"),
            {"id": ref_id, "tid": tid},
        )
    return {"ok": True}


# ─── BIDS HISTORY ─────────────────────────────────────────────────────────────

@router.get("/bids")
def list_bids(user: AuthUser) -> dict[str, Any]:
    engine = get_engine()
    tid = _tenant(user)
    with engine.connect() as conn:
        rows = conn.execute(
            sa.text("""SELECT id, tender_id, cpv, region, our_price, winning_price,
                              n_competitors, won, markup_pct, actual_cost, margin_pct, bid_date
                       FROM historical_bids WHERE org_id=:tid
                       ORDER BY bid_date DESC NULLS LAST"""),
            {"tid": tid},
        ).mappings().fetchall()
    return {"items": [dict(r) for r in rows], "total": len(rows)}


@router.get("/bids/stats")
def bid_stats(user: AuthUser) -> dict[str, Any]:
    engine = get_engine()
    tid = _tenant(user)
    with engine.connect() as conn:
        row = conn.execute(sa.text("""
            SELECT
                COUNT(*) total,
                SUM(CASE WHEN won THEN 1 ELSE 0 END) wins,
                AVG(markup_pct) avg_markup,
                AVG(margin_pct) avg_margin,
                MIN(our_price) min_price,
                MAX(our_price) max_price,
                AVG(CASE WHEN winning_price > 0
                    THEN our_price / winning_price ELSE NULL END) avg_price_ratio
            FROM historical_bids WHERE org_id=:tid
        """), {"tid": tid}).fetchone()

    if not row or not row[0]:
        return {"total": 0, "win_rate_pct": 0, "avg_markup_pct": 0}

    total = row[0]
    return {
        "total":          total,
        "wins":           row[1] or 0,
        "win_rate_pct":   round((row[1] or 0) / total * 100, 1),
        "avg_markup_pct": round(float(row[2] or 0), 1),
        "avg_margin_pct": round(float(row[3] or 0), 1) if row[3] else None,
        "min_price":      float(row[4] or 0),
        "max_price":      float(row[5] or 0),
        "avg_price_ratio": round(float(row[6] or 1), 3) if row[6] else None,
    }


@router.post("/bids")
def add_bid(data: BidIn, user: AuthUser) -> dict[str, Any]:
    engine = get_engine()
    tid = _tenant(user)
    bid_id = str(uuid.uuid4())
    bid_date = date.fromisoformat(data.bid_date) if data.bid_date else date.today()
    with engine.begin() as conn:
        conn.execute(sa.text("""
            INSERT INTO historical_bids
                (id, org_id, tender_id, cpv, region, our_price, winning_price,
                 n_competitors, won, markup_pct, actual_cost, margin_pct, bid_date)
            VALUES (:id, :tid, :tender_id, :cpv, :region, :our_price, :winning_price,
                    :n_competitors, :won, :markup_pct, :actual_cost, :margin_pct, :bid_date)
        """), {
            "id": bid_id, "tid": tid,
            "tender_id": data.tender_id,
            "cpv": data.cpv, "region": data.region,
            "our_price": data.our_price,
            "winning_price": data.winning_price,
            "n_competitors": data.n_competitors,
            "won": data.won,
            "markup_pct": data.markup_pct,
            "actual_cost": data.actual_cost,
            "margin_pct": data.margin_pct,
            "bid_date": bid_date,
        })
    return {"id": bid_id, "ok": True}
